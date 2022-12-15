from openpyxl import load_workbook


wb = load_workbook(filename = 'resourses\Gabarito.xlsx')
ws = wb.active

print(ws["C1"].value)
#for A, B in zip(ws['A'], ws['B']):
#    print(A.value, " - ", B.value)