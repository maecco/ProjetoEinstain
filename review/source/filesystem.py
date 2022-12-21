from os import path

import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.styles import Font


def load_gabarito(PATH):
    wb = xl.load_workbook(filename = PATH)
    ws = wb.active

    gabarito_nao_formatado = list(zip(ws['A'], ws['B']))
    
    # transforma o gabarito de um objeto openpyxl em um dicioanario
    gabarito_formatado = dict()
    for i in range(1, len(gabarito_nao_formatado)):
        linha_do_gabarito = gabarito_nao_formatado[i]
        gabarito_formatado[int(linha_do_gabarito[0].value)] = linha_do_gabarito[1].value
    
    return gabarito_formatado


def load_respostas(PATH):
    wb = xl.load_workbook(filename = PATH)
    ws = wb.active

    respostas_nao_formatadas = list(zip(ws['A'], ws['B'], ws['C']))

    respostas_formatadas = dict()
    
    for i in range(1, len(respostas_nao_formatadas)):
        
        nome_do_aluno = respostas_nao_formatadas[i][0].value
        numero_da_questao = int(respostas_nao_formatadas[i][1].value)
        resposta_do_aluno = respostas_nao_formatadas[i][2].value

        if nome_do_aluno not in respostas_formatadas:
            respostas_formatadas.update({nome_do_aluno:{numero_da_questao:resposta_do_aluno}})
        else:
            respostas_formatadas[nome_do_aluno].update({numero_da_questao:resposta_do_aluno})

    return respostas_formatadas


def save_file(filepath, data):
  
    dirpath = path.dirname(filepath)
    if not path.exists(dirpath):
        return 'path_error'

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Relatorio da correção"

    ws1['A1'] = 'Individual'
    ws1['A1'].font = Font(bold=True)
    ws1['A2'] = 'Aluno'
    ws1['B2'] = 'Acerto prcentual'
    ws1['C2'] = 'Acerto absoluto'

    linha = 3
    for aluno in data['individual']:

        slinha = str(linha)
        ws1['{}'.format("A"+slinha)] = aluno[2]
        ws1['{}'.format("B"+slinha)] = aluno[1]
        ws1['{}'.format("C"+slinha)] = aluno[0]

        ws1['{}'.format("B"+slinha)].number_format = FORMAT_PERCENTAGE_00

        linha += 1


    ws1['E1'] = 'Total'
    ws1['E1'].font = Font(bold=True)
    ws1['E2'] = 'Acerto percentual medio'
    ws1['E3'] = 'Acerto absoluto medio'
    ws1['E4'] = 'Acertos efetivos (sem as provas zeradas)'
    ws1['E4'].font = Font(bold=True)
    ws1['E5'] = 'Provas zeradas'
    ws1['E6'] = 'Acerto percentual efetivo'
    ws1['E7'] = 'Acerto absoluto efetivo'

    total = data['geral']
    ws1['F2'] = total['acerto_percentual']
    ws1['F3'] = total['acerto_absoluto']
    ws1['F5'] = total['provas_zeradas']
    ws1['F6'] = total['acerto_percentual_efetivo']
    ws1['F7'] = total['acerto_absoluto_efetivo']

    ws1['F2'].number_format = FORMAT_PERCENTAGE_00
    ws1['F6'].number_format = FORMAT_PERCENTAGE_00

    wb.save(filename = filepath)
    return 'success'